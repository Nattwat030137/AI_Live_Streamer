"""ตรวจสอบโครงสร้างไฟล์ Excel สินค้า."""

from pathlib import Path

from openpyxl import load_workbook

from config.settings import PROJECT_ROOT


EXCEL_FILE = PROJECT_ROOT / "data" / "products_source.xlsx"


def main() -> None:
    """แสดงชื่อชีต ขนาดตาราง และข้อมูลแถวแรก ๆ."""

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Excel: {EXCEL_FILE}"
        )

    workbook = load_workbook(
        filename=EXCEL_FILE,
        read_only=True,
        data_only=True,
    )

    print("=" * 60)
    print(f"ไฟล์: {EXCEL_FILE.name}")
    print(f"รายชื่อชีต: {workbook.sheetnames}")
    print("=" * 60)

    worksheet = workbook[workbook.sheetnames[0]]

    print(f"ชีตที่กำลังตรวจ: {worksheet.title}")
    print(f"จำนวนแถวทั้งหมด: {worksheet.max_row}")
    print(f"จำนวนคอลัมน์ทั้งหมด: {worksheet.max_column}")
    print("=" * 60)

    print("ข้อมูล 6 แถวแรก:")

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=6,
            values_only=True,
        ),
        start=1,
    ):
        print(f"แถว {row_number}: {row}")

    workbook.close()


if __name__ == "__main__":
    main()