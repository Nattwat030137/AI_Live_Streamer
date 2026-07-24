"""Regression tests for the Excel product catalog importer."""

import json
import sqlite3

from pathlib import Path

from openpyxl import (
    Workbook,
    load_workbook,
)

from app.catalog.init_product_db import (
    initialize_database,
    load_excel_products,
)


def test_load_excel_products_deduplicates_models(
    tmp_path: Path,
) -> None:
    """Use SKU models, fall back to names, and merge duplicates."""

    excel_path = tmp_path / "products.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet 1"

    worksheet.append(
        ["Export Products Link"]
    )
    worksheet.append(
        ["Period", "Test"]
    )
    worksheet.append(
        [
            "Connection name",
            "Channel",
            "Store",
            "Product name",
            "SKU",
            "Sold",
            "Description",
            "Product type",
            "Updated",
            "Creator",
            "Created",
        ]
    )
    worksheet.append(
        [
            "Canonical Product",
            "Shopee",
            "Store A",
            "Marketplace title A",
            "SKU-001",
            3,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )
    worksheet.append(
        [
            "Canonical Product",
            "Lazada",
            "Store B",
            "Marketplace title B",
            "SKU-001",
            3,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )
    worksheet.append(
        [
            "Fallback Model",
            "Manual Product",
            "Store A",
            "Fallback product title",
            "",
            1,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )

    workbook.save(excel_path)
    workbook.close()

    products = load_excel_products(
        excel_path
    )

    products_by_model = {
        product["model"]: product
        for product in products
    }

    assert len(products) == 2
    assert set(products_by_model) == {
        "SKU-001",
        "Fallback Model",
    }
    assert (
        products_by_model["SKU-001"]["name"]
        == "Canonical Product"
    )
    assert (
        products_by_model["SKU-001"]["source"]
        == "products.xlsx"
    )


def test_initialize_database_merges_excel_safely(
    tmp_path: Path,
) -> None:
    """Import Excel products without replacing curated data."""

    json_path = tmp_path / "curated_catalog.json"
    excel_path = tmp_path / "products.xlsx"
    database_path = tmp_path / "products.db"

    json_path.write_text(
        json.dumps(
            [
                {
                    "model": "5073",
                    "name": "Curated Tray 5073",
                    "category": "Tray",
                    "compatible_bag": "12 x 20 cm",
                    "material": "PET",
                    "color": "Clear",
                    "notes": "Curated product data",
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet 1"

    worksheet.append(
        ["Export Products Link"]
    )
    worksheet.append(
        ["Period", "Test"]
    )
    worksheet.append(
        [
            "Connection name",
            "Channel",
            "Store",
            "Product name",
            "SKU",
            "Sold",
            "Description",
            "Product type",
            "Updated",
            "Creator",
            "Created",
        ]
    )
    worksheet.append(
        [
            "Excel Tray 5073",
            "Shopee",
            "Store A",
            "Marketplace tray title",
            "5073",
            5,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )
    worksheet.append(
        [
            "Excel Product",
            "Shopee",
            "Store A",
            "Marketplace product title",
            "SKU-001",
            2,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )
    worksheet.append(
        [
            "Fallback Model",
            "Manual Product",
            "Store A",
            "Fallback product title",
            "",
            1,
            "-",
            "Standard",
            "",
            "",
            "",
        ]
    )

    workbook.save(excel_path)
    workbook.close()

    imported_count = initialize_database(
        json_path=json_path,
        excel_path=excel_path,
        database_path=database_path,
    )

    with sqlite3.connect(
        database_path
    ) as connection:
        total = connection.execute(
            "SELECT COUNT(*) FROM products"
        ).fetchone()[0]

        curated_row = connection.execute(
            """
            SELECT
                name,
                compatible_bag,
                material,
                source
            FROM products
            WHERE model = ?
            """,
            ("5073",),
        ).fetchone()

    assert imported_count == 3
    assert total == 3
    assert curated_row == (
        "Curated Tray 5073",
        "12 x 20 cm",
        "PET",
        "curated_catalog.json",
    )
    assert curated_row == (
        "Curated Tray 5073",
        "12 x 20 cm",
        "PET",
        "curated_catalog.json",
    )
    workbook = load_workbook(
        excel_path
    )
    worksheet = workbook[
        "Sheet 1"
    ]
    worksheet.delete_rows(
        6
    )
    workbook.save(
        excel_path
    )
    workbook.close()

    rebuilt_count = initialize_database(
        json_path=json_path,
        excel_path=excel_path,
        database_path=database_path,
    )

    with sqlite3.connect(
        database_path
    ) as connection:
        stale_row = connection.execute(
            """
            SELECT model
            FROM products
            WHERE model = ?
            """,
            ("Fallback Model",),
        ).fetchone()

    assert rebuilt_count == 2
    assert stale_row is None


def test_real_catalog_builds_expected_total(
    tmp_path: Path,
) -> None:
    """Build the complete tracked catalog in a temporary database."""

    project_root = (
        Path(__file__)
        .resolve()
        .parents[1]
    )
    json_path = (
        project_root
        / "data"
        / "bakery_products.json"
    )
    excel_path = (
        project_root
        / "data"
        / "products_source.xlsx"
    )
    database_path = (
        tmp_path
        / "products.db"
    )

    imported_count = initialize_database(
        json_path=json_path,
        excel_path=excel_path,
        database_path=database_path,
    )

    with sqlite3.connect(
        database_path
    ) as connection:
        excel_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM products
            WHERE source = ?
            """,
            ("products_source.xlsx",),
        ).fetchone()[0]

        curated_rows = connection.execute(
            """
            SELECT
                model,
                material,
                compatible_bag,
                source
            FROM products
            WHERE model IN (?, ?)
            ORDER BY model
            """,
            (
                "5040",
                "5073",
            ),
        ).fetchall()

    assert imported_count == 1235
    assert excel_count == 1233
    assert curated_rows == [
        (
            "5040",
            "กระดาษ",
            "",
            "bakery_products.json",
        ),
        (
            "5073",
            "PET",
            "12 x 20 ซม.",
            "bakery_products.json",
        ),
    ]


def test_legacy_importer_targets_runtime_database() -> None:
    """Keep the legacy command pointed at the runtime catalog."""

    from app.product_importer import (
        DATABASE_FILE,
        EXCEL_FILE,
    )

    project_root = (
        Path(__file__)
        .resolve()
        .parents[1]
    )

    assert EXCEL_FILE == (
        project_root
        / "data"
        / "products_source.xlsx"
    )
    assert DATABASE_FILE == (
        project_root
        / "data"
        / "products.db"
    )
