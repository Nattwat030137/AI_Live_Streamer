"""สร้าง SQLite Product Catalog จากไฟล์ JSON."""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DEFAULT_JSON_PATH = (
    PROJECT_ROOT
    / "data"
    / "bakery_products.json"
)

DEFAULT_EXCEL_PATH = (
    PROJECT_ROOT
    / "data"
    / "products_source.xlsx"
)

DEFAULT_DATABASE_PATH = (
    PROJECT_ROOT
    / "data"
    / "products.db"
)


CREATE_PRODUCTS_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    model TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    category TEXT NOT NULL DEFAULT '',
    compatible_bag TEXT NOT NULL DEFAULT '',
    material TEXT NOT NULL DEFAULT '',
    color TEXT NOT NULL DEFAULT '',
    notes TEXT NOT NULL DEFAULT '',
    source TEXT NOT NULL DEFAULT 'bakery_products.json',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""


CREATE_MODEL_INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_products_model
ON products(model);
"""


CREATE_CATEGORY_INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_products_category
ON products(category);
"""


UPSERT_PRODUCT_SQL = """
INSERT INTO products (
    model,
    name,
    category,
    compatible_bag,
    material,
    color,
    notes,
    source
)
VALUES (?, ?, ?, ?, ?, ?, ?, ?)
ON CONFLICT(model) DO UPDATE SET
    name = excluded.name,
    category = excluded.category,
    compatible_bag = excluded.compatible_bag,
    material = excluded.material,
    color = excluded.color,
    notes = excluded.notes,
    source = excluded.source,
    updated_at = CURRENT_TIMESTAMP;
"""
INSERT_EXCEL_PRODUCT_SQL = """
INSERT INTO products (
    model,
    name,
    category,
    compatible_bag,
    material,
    color,
    notes,
    source
)
VALUES (?, ?, ?, ?, ?, ?, ?, ?)
ON CONFLICT(model) DO NOTHING;
"""


def load_json_products(
    json_path: Path,
) -> list[dict[str, Any]]:
    """โหลดและตรวจข้อมูลสินค้าจาก JSON."""

    if not json_path.exists():
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Product Catalog: {json_path}"
        )

    raw_text = json_path.read_text(
        encoding="utf-8"
    )

    raw_data = json.loads(
        raw_text
    )

    if not isinstance(
        raw_data,
        list,
    ):
        raise ValueError(
            "Product Catalog ต้องเป็น JSON List"
        )

    products: list[dict[str, Any]] = []

    for index, item in enumerate(
        raw_data,
        start=1,
    ):
        if not isinstance(
            item,
            dict,
        ):
            raise ValueError(
                f"ข้อมูลสินค้าลำดับที่ {index} "
                "ต้องเป็น JSON Object"
            )

        model = str(
            item.get(
                "model",
                "",
            )
        ).strip()

        name = str(
            item.get(
                "name",
                "",
            )
        ).strip()

        if not model:
            raise ValueError(
                f"สินค้าลำดับที่ {index} ไม่มี model"
            )

        if not name:
            raise ValueError(
                f"สินค้าลำดับที่ {index} ไม่มี name"
            )

        products.append(
            {
                "model": model,
                "name": name,
                "category": str(
                    item.get(
                        "category",
                        "",
                    )
                ).strip(),
                "compatible_bag": str(
                    item.get(
                        "compatible_bag",
                        "",
                    )
                ).strip(),
                "material": str(
                    item.get(
                        "material",
                        "",
                    )
                ).strip(),
                "color": str(
                    item.get(
                        "color",
                        "",
                    )
                ).strip(),
                "notes": str(
                    item.get(
                        "notes",
                        "",
                    )
                ).strip(),
                "source": json_path.name,
            }
        )

    return products


def _clean_excel_text(
    value: Any,
) -> str:
    """Convert an Excel value to clean text."""

    if value is None:
        return ""

    return str(value).strip()


def load_excel_products(
    excel_path: Path,
    *,
    sheet_name: str = "Sheet 1",
    first_data_row: int = 4,
) -> list[dict[str, Any]]:
    """Load and deduplicate products exported from GoSell."""

    if not excel_path.exists():
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Product Excel: {excel_path}"
        )

    workbook = load_workbook(
        filename=excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"ไม่พบชีต {sheet_name} "
                f"ชีตที่มีอยู่คือ {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        products_by_model: dict[
            str,
            dict[str, Any],
        ] = {}

        for row in worksheet.iter_rows(
            min_row=first_data_row,
            values_only=True,
        ):
            connection_name = (
                _clean_excel_text(row[0])
            )
            product_name = (
                _clean_excel_text(row[3])
            )
            sku = _clean_excel_text(
                row[4]
            )
            description = (
                _clean_excel_text(row[6])
            )
            category = (
                _clean_excel_text(row[7])
            )

            model = (
                sku
                or connection_name
            )
            name = (
                connection_name
                or product_name
                or model
            )

            if not model or not name:
                continue

            notes_parts = [
                value
                for value in (
                    product_name,
                    description,
                )
                if (
                    value
                    and value != "-"
                    and value != name
                )
            ]
            notes = " | ".join(
                notes_parts
            )

            normalized_model = (
                model.casefold()
            )
            existing_product = (
                products_by_model.get(
                    normalized_model
                )
            )

            if existing_product is None:
                products_by_model[
                    normalized_model
                ] = {
                    "model": model,
                    "name": name,
                    "category": category,
                    "compatible_bag": "",
                    "material": "",
                    "color": "",
                    "notes": notes,
                    "source": excel_path.name,
                }
                continue

            if (
                not existing_product["category"]
                and category
            ):
                existing_product[
                    "category"
                ] = category

            if (
                notes
                and notes
                not in existing_product["notes"]
            ):
                existing_notes = (
                    existing_product["notes"]
                )
                existing_product["notes"] = (
                    f"{existing_notes} | {notes}"
                    if existing_notes
                    else notes
                )

        return list(
            products_by_model.values()
        )

    finally:
        workbook.close()


def initialize_database(
    *,
    json_path: Path = DEFAULT_JSON_PATH,
    excel_path: Path | None = (
        DEFAULT_EXCEL_PATH
    ),
    database_path: Path = DEFAULT_DATABASE_PATH,
) -> int:
    """
    Build the runtime catalog from curated JSON and Excel.

    Curated JSON products take precedence when models overlap.

    Returns:
        Total number of products in the database.
    """

    curated_products = load_json_products(
        json_path
    )

    excel_products = (
        load_excel_products(
            excel_path
        )
        if excel_path is not None
        else []
    )

    curated_models = {
        product["model"].casefold()
        for product in curated_products
    }

    database_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with sqlite3.connect(
        database_path
    ) as connection:
        connection.execute(
            CREATE_PRODUCTS_TABLE_SQL
        )

        connection.execute(
            CREATE_MODEL_INDEX_SQL
        )

        connection.execute(
            CREATE_CATEGORY_INDEX_SQL
        )

        connection.execute(
            """
            DELETE FROM products
            WHERE source = ?
            """,
            (
                json_path.name,
            ),
        )

        if excel_path is not None:
            connection.execute(
                """
                DELETE FROM products
                WHERE source = ?
                """,
                (
                    excel_path.name,
                ),
            )

        for product in curated_products:
            connection.execute(
                UPSERT_PRODUCT_SQL,
                (
                    product["model"],
                    product["name"],
                    product["category"],
                    product["compatible_bag"],
                    product["material"],
                    product["color"],
                    product["notes"],
                    product["source"],
                ),
            )

        for product in excel_products:
            if (
                product["model"].casefold()
                in curated_models
            ):
                continue

            connection.execute(
                INSERT_EXCEL_PRODUCT_SQL,
                (
                    product["model"],
                    product["name"],
                    product["category"],
                    product["compatible_bag"],
                    product["material"],
                    product["color"],
                    product["notes"],
                    product["source"],
                ),
            )

        connection.commit()

        count_row = connection.execute(
            """
            SELECT COUNT(*)
            FROM products
            """
        ).fetchone()

    if count_row is None:
        return 0

    return int(
        count_row[0]
    )


def read_database_summary(
    database_path: Path = DEFAULT_DATABASE_PATH,
) -> list[tuple[str, str]]:
    """อ่าน Model และชื่อสินค้าจากฐานข้อมูล."""

    if not database_path.exists():
        raise FileNotFoundError(
            f"ไม่พบฐานข้อมูล: {database_path}"
        )

    with sqlite3.connect(
        database_path
    ) as connection:
        rows = connection.execute(
            """
            SELECT model, name
            FROM products
            ORDER BY model
            """
        ).fetchall()

    return [
        (
            str(row[0]),
            str(row[1]),
        )
        for row in rows
    ]


def main() -> None:
    """สร้างฐานข้อมูลและแสดงผลสรุป."""

    imported_count = initialize_database()

    rows = read_database_summary()

    print("=" * 60)
    print("SQLite Product Catalog")
    print("=" * 60)
    print(
        f"Database: {DEFAULT_DATABASE_PATH}"
    )
    print(
        f"Imported: {imported_count}"
    )
    print(
        f"Products in database: {len(rows)}"
    )
    print("-" * 60)

    for model, name in rows:
        print(
            f"{model}: {name}"
        )

    print("=" * 60)


if __name__ == "__main__":
    main()