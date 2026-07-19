"""นำเข้าข้อมูลสินค้าจาก Excel ไปยังฐานข้อมูล SQLite."""

import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.settings import PROJECT_ROOT


EXCEL_FILE = PROJECT_ROOT / "data" / "products_source.xlsx"
DATABASE_FILE = PROJECT_ROOT / "database" / "products.db"

SHEET_NAME = "Sheet 1"
HEADER_ROW = 3
FIRST_DATA_ROW = 4


def clean_text(value: Any) -> str:
    """เปลี่ยนข้อมูลจาก Excel ให้เป็นข้อความที่สะอาด."""

    if value is None:
        return ""

    return str(value).strip()


def clean_integer(value: Any) -> int:
    """เปลี่ยนข้อมูลเป็นจำนวนเต็ม หากเปลี่ยนไม่ได้ให้ใช้ศูนย์."""

    if value is None:
        return 0

    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def create_database(connection: sqlite3.Connection) -> None:
    """สร้างตารางสินค้า หากยังไม่มี."""

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS product_listings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            connection_name TEXT NOT NULL,
            channel TEXT NOT NULL,
            channel_store_name TEXT,
            product_name TEXT NOT NULL,
            sku TEXT,
            sold_count INTEGER DEFAULT 0,
            description TEXT,
            product_type TEXT,
            updated_at TEXT,
            created_by TEXT,
            created_at TEXT,
            source_row INTEGER,
            UNIQUE (
                connection_name,
                channel,
                channel_store_name,
                product_name
            )
        )
        """
    )

    connection.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_product_name
        ON product_listings(product_name)
        """
    )

    connection.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_product_sku
        ON product_listings(sku)
        """
    )

    connection.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_product_channel
        ON product_listings(channel)
        """
    )

    connection.commit()


def import_products() -> int:
    """อ่าน Excel และนำสินค้าเข้าฐานข้อมูล."""

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Excel: {EXCEL_FILE}"
        )

    DATABASE_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = load_workbook(
        filename=EXCEL_FILE,
        read_only=True,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"ไม่พบชีต {SHEET_NAME} "
            f"ชีตที่มีอยู่คือ {workbook.sheetnames}"
        )

    worksheet = workbook[SHEET_NAME]

    connection = sqlite3.connect(DATABASE_FILE)

    imported_count = 0

    try:
        create_database(connection)

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=FIRST_DATA_ROW,
                values_only=True,
            ),
            start=FIRST_DATA_ROW,
        ):
            connection_name = clean_text(row[0])
            channel = clean_text(row[1])
            channel_store_name = clean_text(row[2])
            product_name = clean_text(row[3])
            sku = clean_text(row[4])
            sold_count = clean_integer(row[5])
            description = clean_text(row[6])
            product_type = clean_text(row[7])
            updated_at = clean_text(row[8])
            created_by = clean_text(row[9])
            created_at = clean_text(row[10])

            # ข้ามแถวว่าง
            if not product_name and not sku and not connection_name:
                continue

            connection.execute(
                """
                INSERT INTO product_listings (
                    connection_name,
                    channel,
                    channel_store_name,
                    product_name,
                    sku,
                    sold_count,
                    description,
                    product_type,
                    updated_at,
                    created_by,
                    created_at,
                    source_row
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (
                    connection_name,
                    channel,
                    channel_store_name,
                    product_name
                )
                DO UPDATE SET
                    sku = excluded.sku,
                    sold_count = excluded.sold_count,
                    description = excluded.description,
                    product_type = excluded.product_type,
                    updated_at = excluded.updated_at,
                    created_by = excluded.created_by,
                    created_at = excluded.created_at,
                    source_row = excluded.source_row
                """,
                (
                    connection_name,
                    channel,
                    channel_store_name,
                    product_name,
                    sku,
                    sold_count,
                    description,
                    product_type,
                    updated_at,
                    created_by,
                    created_at,
                    row_number,
                ),
            )

            imported_count += 1

        connection.commit()

    finally:
        connection.close()
        workbook.close()

    return imported_count


def main() -> None:
    """เริ่มกระบวนการนำเข้าสินค้า."""

    print("กำลังนำเข้าสินค้าจาก Excel...")

    imported_count = import_products()

    print("=" * 60)
    print("นำเข้าสินค้าสำเร็จ")
    print(f"จำนวนแถวที่ประมวลผล: {imported_count:,}")
    print(f"ฐานข้อมูล: {DATABASE_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()